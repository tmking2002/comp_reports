import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.styles import Alignment
import streamlit as st
from dotenv import load_dotenv
import os
import pymysql
from tempfile import NamedTemporaryFile
from io import BytesIO

load_dotenv()

connection = pymysql.connect(
    host=os.getenv("DB_HOST"),
    user=os.getenv("DB_USERNAME"),
    password=os.getenv("DB_PASSWORD"),
    database=os.getenv("DB_NAME")
)

cursor = connection.cursor()

# Get list of all players 
query = """
    SELECT Player
    FROM (
        SELECT CONCAT(first_name, ' ', last_name, ' (', ncaa_university_name, ')') as Player
        FROM diamond_position_full
        WHERE cycle_id = 6 AND first_name NOT REGEXP '[0-9.]'

        UNION

        SELECT CONCAT(first_name, ' ', last_name, ' (', ncaa_university_name, ')') as Player
        FROM diamond_pitching_full
        WHERE cycle_id = 6 AND first_name NOT REGEXP '[0-9.]'
    ) AS subquery
    ORDER BY SUBSTRING_INDEX(Player, ' ', 1);
"""

cursor.execute(query)

rows = cursor.fetchall()
columns = [desc[0] for desc in cursor.description]
all_players = pd.DataFrame(rows, columns=columns)


query = """
    SELECT ncaa_university_link.ncaa_university_name, ncaa_university_link.ncaa_universityID
    FROM ncaa_university_link
"""

cursor.execute(query)

rows = cursor.fetchall()
columns = [desc[0] for desc in cursor.description]
teams = pd.DataFrame(rows, columns=columns)

def create_sheet(team_id, transfers_in, players_leaving):
    # Create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Merge cells A1 to K1
    sheet.merge_cells('A1:K1')
    font = Font(size=36, bold=True)
    for col in range(1, 26):
        cell = sheet.cell(row=1, column=col)
        cell.font = font

    # Set the value of the merged cells to 'Hitting Stats' and center align the text
    sheet['A1'] = 'Hitting Stats'
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Add border to cells A1 to K1
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, 12):
        cell = sheet.cell(row=1, column=col)
        cell.border = border

    # Get team name

    query = """
        SELECT ncaa_university_name
        FROM ncaa_university_link
        WHERE ncaa_universityID = {}
    """.format(team_id)

    cursor.execute(query)

    rows = cursor.fetchall()
    team_name = rows[0][0]

    # Get name of conference and conference_id

    query = """
        SELECT university_conference.ncaa_universityID, conference.conference, conference.conference_id
        FROM university_conference
        LEFT JOIN conference on university_conference.conference_id = conference.conference_id
        WHERE university_conference.ncaa_universityID = {}
    """.format(team_id)

    cursor.execute(query)

    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    conference = pd.DataFrame(rows, columns=columns)

    # Set the value of merged cells E4:G4 to the conference name
    sheet.merge_cells('D4:G4')
    sheet['D4'] = conference['conference'][0]
    sheet['D4'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet['D4'].font = Font(size=16, bold=True)

    sheet.merge_cells('D5:E5')
    sheet['D5'] = 'Team'
    sheet['D5'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet['D5'].font = Font(bold = True)

    sheet['F5'] = 'wRAA'
    sheet['F5'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet['F5'].font = Font(bold = True)

    sheet['G5'] = 'Record'
    sheet['G5'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet['G5'].font = Font(bold = True)

    # Get conference teams

    query = """
        SELECT ncaa_university_link.ncaa_university_name, ncaa_university_link.ncaa_universityID
        FROM university_conference
        LEFT JOIN ncaa_university_link ON ncaa_university_link.ncaa_universityID = university_conference.ncaa_universityID
        WHERE university_conference.conference_id = {}
    """.format(conference['conference_id'][0])

    cursor.execute(query)

    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    conference_teams = pd.DataFrame(rows, columns=columns)

    # Get hitting stats for conference teams

    query = """
        SELECT ncaa_university_link.ncaa_university_name, SUM(wRAA) AS wRAA
        FROM diamond_position_full
        LEFT JOIN ncaa_university_link ON ncaa_university_link.ncaa_university_name = diamond_position_full.ncaa_university_name
        WHERE ncaa_university_link.ncaa_university_name IN {} AND cycle_id = 6
        GROUP BY ncaa_university_link.ncaa_university_name
        ORDER BY wRAA DESC
    """.format(tuple(conference_teams['ncaa_university_name'].tolist()))

    cursor.execute(query)

    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    conference_hitting_stats = pd.DataFrame(rows, columns=columns)

    # Get conference records
    query = """
        SELECT ncaa_college_game_id, game_date, home_team_id, away_team_id, CASE WHEN home_team_runs > away_team_runs THEN 1 ELSE 0 END as home_win
        FROM ncaa_college_games
        WHERE (home_team_id IN {}) AND (away_team_id IN {}) AND (season = 2023)
        ORDER BY game_date
    """.format(tuple(conference_teams['ncaa_universityID'].tolist()), tuple(conference_teams['ncaa_universityID'].tolist()))

    cursor.execute(query)

    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    conference_games = pd.DataFrame(rows, columns=columns)

    conference_records = pd.DataFrame(columns=['team_id', 'record'])

    for team in conference_teams['ncaa_universityID'].tolist():
        team_home_games = conference_games[conference_games['home_team_id'] == team]

        home_wins = team_home_games[team_home_games['home_win'] == 1].shape[0]
        home_losses = len(team_home_games) - home_wins

        team_away_games = conference_games[conference_games['away_team_id'] == team]

        away_losses = team_away_games[team_away_games['home_win'] == 1].shape[0]
        away_wins = len(team_away_games) - away_losses

        total_wins = home_wins + away_wins
        total_losses = home_losses + away_losses

        new_row = {'team_id': team, 'record': f'{total_wins}-{total_losses}'}
        conference_records = pd.concat([conference_records, pd.DataFrame([new_row])], ignore_index=True)

    conference_records = pd.merge(conference_records, conference_teams, left_on="team_id", right_on='ncaa_universityID')

    conference_hitting_stats = pd.merge(conference_hitting_stats, conference_records, left_on='ncaa_university_name', right_on="ncaa_university_name")

    for index, row in conference_hitting_stats.iterrows():
        sheet['D{}'.format(index + 6)] = row['ncaa_university_name']
        sheet['D{}'.format(index + 6)].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet.merge_cells('D{}:E{}'.format(index + 6, index + 6))
        sheet['F{}'.format(index + 6)] = round(row['wRAA'], 2)
        sheet['F{}'.format(index + 6)].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet['G{}'.format(index + 6)] = row['record']
        sheet['G{}'.format(index + 6)].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    for col in range(4, 8):
        for row in range(4, len(conference_hitting_stats) + 6):
            cell = sheet.cell(row=row, column=col)
            if row == 4 and col == 4:
                cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'), left=openpyxl.styles.Side(style='thin'))
            elif row == 4 and col == 7:
                cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
            elif row == len(conference_hitting_stats) + 5 and col == 4:
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), left=openpyxl.styles.Side(style='thin'))
            elif row == len(conference_hitting_stats) + 5 and col == 7:
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
            elif row == 4:
                cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'))
            elif row == len(conference_hitting_stats) + 5:
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
            elif col == 4:
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'))
            elif col == 7:
                cell.border = openpyxl.styles.Border(right=openpyxl.styles.Side(style='thin'))
            
            if cell.value == team_name:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFF157', end_color='FFF157', fill_type='solid')
                

    # Extract player and team information
    players = [entry.split(' (')[0] for entry in transfers_in]
    teams = [entry.split(' (')[1].replace(')', '') for entry in transfers_in]

    # Create a DataFrame
    transfers_in_df = pd.DataFrame({'Player': players, 'Team': teams})

    # Get the hitting stats for the selected team

    query = """
        SELECT diamond_position_full.ncaa_university_name, concat(first_name, ' ', last_name) as Name, Yr, Pos, PA, H, HR, BA, SlgPct as SLG, OBPct + SlgPct as OPS, wRAA, wRAA / PA * 100 as wRAA_per_100
        FROM diamond_position_full
        LEFT JOIN ncaa_university_link on ncaa_university_link.ncaa_university_name = diamond_position_full.ncaa_university_name
        WHERE cycle_id = 6 AND ncaa_university_link.ncaa_universityID = {} AND PA > 0
    """.format(team_id)

    cursor.execute(query)

    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]  
    hitting_stats = pd.DataFrame(rows, columns=columns)  

    if len(transfers_in_df) > 0:
        query = """
            SELECT diamond_position_full.ncaa_university_name, concat(first_name, ' ', last_name) as Name, Yr, Pos, PA, H, HR, BA, SlgPct as SLG, OBPct + SlgPct as OPS, wRAA, wRAA / PA * 100 as wRAA_per_100
            FROM diamond_position_full
            LEFT JOIN ncaa_university_link on ncaa_university_link.ncaa_university_name = diamond_position_full.ncaa_university_name
            WHERE cycle_id = 6 AND Name IN {} AND PA > 0
        """.format(tuple(transfers_in_df['Player'].tolist()))

        cursor.execute(query)

        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        transfers_in_stats = pd.DataFrame(rows, columns=columns)
    else:
        transfers_in_stats = pd.DataFrame()


    hitting_stats = pd.concat([hitting_stats, transfers_in_stats], ignore_index=True)
    hitting_stats = pd.merge(hitting_stats, transfers_in_df, left_on=['Name', 'ncaa_university_name'], right_on=['Player', 'Team'], how='left')
    hitting_stats = hitting_stats.drop(columns=['Team', 'Player'])

    hitting_stats['BA'] = hitting_stats['BA'].astype(float).round(3).apply('{:.3f}'.format)
    hitting_stats['SLG'] = hitting_stats['SLG'].astype(float).round(3).apply('{:.3f}'.format)
    hitting_stats['OPS'] = hitting_stats['OPS'].astype(float).round(3).apply('{:.3f}'.format)

    hitting_stats['wRAA'] = hitting_stats['wRAA'].astype(float).round(2)
    hitting_stats['wRAA_per_100'] = hitting_stats['wRAA_per_100'].astype(float).round(2)

    hitting_stats = hitting_stats.sort_values(by=['wRAA_per_100'], ascending=False)
    hitting_stats = hitting_stats.reset_index(drop=True)

    for index, colname in enumerate(['Team', 'Player', 'Year', 'Pos', 'PA', 'H', 'HR', 'BA', 'SLG', 'OPS', 'wRAA', 'wRAA/100']):
        cell = sheet.cell(row=len(conference_hitting_stats) + 8, column=index + 1)
        cell.value = colname
        cell.font = Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


    for index, row in hitting_stats.iterrows():
        for col in range(1, 13):
            cell = sheet.cell(row=index + len(conference_hitting_stats) + 9, column=col)
            cell.value = row[col - 1]
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            if not hitting_stats['ncaa_university_name'][index] == team_name:
                cell.fill = openpyxl.styles.PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
            elif hitting_stats['Name'][index] in players_leaving:
                cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            else:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        

    # Set column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 6
    sheet.column_dimensions['D'].width = 6

    sheet.title = "Hitting"

    returning_color_cell = sheet.cell(row=len(conference_hitting_stats) + len(hitting_stats) + 12, column=1)
    returning_color_cell.value = "Returners"
    returning_color_cell.fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    returning_color_cell.font = Font(bold=True)

    transfer_color_cell = sheet.cell(row=len(conference_hitting_stats) + len(hitting_stats) + 13, column=1)
    transfer_color_cell.value = "Transfers"
    transfer_color_cell.fill = openpyxl.styles.PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    transfer_color_cell.font = Font(bold=True)

    leaving_color_cell = sheet.cell(row=len(conference_hitting_stats) + len(hitting_stats) + 14, column=1)
    leaving_color_cell.value = "Leaving"
    leaving_color_cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    leaving_color_cell.font = Font(bold=True)


    sheet = workbook.create_sheet("Pitching")

    # Merge cells A1 to K1
    sheet.merge_cells('A1:K1')
    font = Font(size=36, bold=True)
    for col in range(1, 26):
        cell = sheet.cell(row=1, column=col)
        cell.font = font

    # Set the value of the merged cells to 'Hitting Stats' and center align the text
    sheet['A1'] = 'Pitching Stats'
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Add border to cells A1 to K1
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, 12):
        cell = sheet.cell(row=1, column=col)
        cell.border = border

    # Get team name

    query = """
        SELECT ncaa_university_name
        FROM ncaa_university_link
        WHERE ncaa_universityID = {}
    """.format(team_id)

    cursor.execute(query)

    rows = cursor.fetchall()
    team_name = rows[0][0]

    # Get name of conference and conference_id

    query = """
        SELECT university_conference.ncaa_universityID, conference.conference, conference.conference_id
        FROM university_conference
        LEFT JOIN conference on university_conference.conference_id = conference.conference_id
        WHERE university_conference.ncaa_universityID = {}
    """.format(team_id)

    cursor.execute(query)

    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    conference = pd.DataFrame(rows, columns=columns)

    # Set the value of merged cells E4:G4 to the conference name
    sheet.merge_cells('D4:G4')
    sheet['D4'] = conference['conference'][0]
    sheet['D4'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet['D4'].font = Font(size=16, bold=True)

    sheet.merge_cells('D5:E5')
    sheet['D5'] = 'Team'
    sheet['D5'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet['D5'].font = Font(bold = True)

    sheet['F5'] = 'FIP'
    sheet['F5'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet['F5'].font = Font(bold = True)

    sheet['G5'] = 'Record'
    sheet['G5'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet['G5'].font = Font(bold = True)

    query = """
        SELECT ncaa_university_link.ncaa_university_name, SUM(FIP * IP) / SUM(IP) as FIP
        FROM diamond_pitching_full
        LEFT JOIN ncaa_university_link ON ncaa_university_link.ncaa_university_name = diamond_pitching_full.ncaa_university_name
        WHERE ncaa_university_link.ncaa_university_name in {} AND cycle_id = 6
        GROUP BY ncaa_university_link.ncaa_university_name
        ORDER BY FIP
    """.format(tuple(conference_teams['ncaa_university_name'].tolist()))

    cursor.execute(query)

    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    conference_pitching_stats = pd.DataFrame(rows, columns=columns)

    conference_pitching_stats = pd.merge(conference_pitching_stats, conference_records, left_on='ncaa_university_name', right_on="ncaa_university_name")

    for index, row in conference_pitching_stats.iterrows():
        sheet['D{}'.format(index + 6)] = row['ncaa_university_name']
        sheet['D{}'.format(index + 6)].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet.merge_cells('D{}:E{}'.format(index + 6, index + 6))
        sheet['F{}'.format(index + 6)] = round(row['FIP'], 2)
        sheet['F{}'.format(index + 6)].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet['G{}'.format(index + 6)] = row['record']
        sheet['G{}'.format(index + 6)].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    for col in range(4, 8):
        for row in range(4, len(conference_pitching_stats) + 6):
            cell = sheet.cell(row=row, column=col)
            if row == 4 and col == 4:
                cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'), left=openpyxl.styles.Side(style='thin'))
            elif row == 4 and col == 7:
                cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
            elif row == len(conference_pitching_stats) + 5 and col == 4:
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), left=openpyxl.styles.Side(style='thin'))
            elif row == len(conference_pitching_stats) + 5 and col == 7:
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
            elif row == 4:
                cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'))
            elif row == len(conference_pitching_stats) + 5:
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
            elif col == 4:
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'))
            elif col == 7:
                cell.border = openpyxl.styles.Border(right=openpyxl.styles.Side(style='thin'))
            
            if cell.value == team_name:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFF157', end_color='FFF157', fill_type='solid')
                
    query = """
        SELECT diamond_pitching_full.ncaa_university_name, concat(first_name, ' ', last_name) as Name, Yr, App, IP, W, L, SO, HA, HR_A, WHIP, HA / (BF - BB - HB - SHA - SFA) AS BAA, FIP
        FROM diamond_pitching_full
        LEFT JOIN ncaa_university_link on ncaa_university_link.ncaa_university_name = diamond_pitching_full.ncaa_university_name
        WHERE cycle_id = 6 AND ncaa_university_link.ncaa_universityID = {} AND BF > 0
        ORDER BY FIP
    """.format(team_id)

    cursor.execute(query)

    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]  
    pitching_stats = pd.DataFrame(rows, columns=columns)  

    if len(transfers_in_df) > 0:
        query = """
            SELECT diamond_pitching_full.ncaa_university_name, concat(first_name, ' ', last_name) as Name, Yr, App, IP, W, L, SO, HA, HR_A, WHIP, HA / (BF - BB - HB - SHA - SFA) AS BAA, FIP
            FROM diamond_pitching_full
            LEFT JOIN ncaa_university_link on ncaa_university_link.ncaa_university_name = diamond_pitching_full.ncaa_university_name
            WHERE cycle_id = 6 AND CONCAT(diamond_pitching_full.first_name, ' ', diamond_pitching_full.last_name) IN {} AND BF > 0
        """.format(tuple(transfers_in_df['Player'].tolist()))

        cursor.execute(query)

        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        transfers_in_stats = pd.DataFrame(rows, columns=columns)
    else:
        transfers_in_stats = pd.DataFrame()

    pitching_stats = pd.concat([pitching_stats, transfers_in_stats], ignore_index=True)
    pitching_stats = pd.merge(pitching_stats, transfers_in_df, left_on=['Name', 'ncaa_university_name'], right_on=['Player', 'Team'], how='left')
    pitching_stats = pitching_stats.drop(columns=['Team', 'Player'])

    pitching_stats['BAA'] = pitching_stats['BAA'].astype(float).round(3).apply('{:.3f}'.format)
    pitching_stats['WHIP'] = pitching_stats['WHIP'].astype(float).round(2)
    pitching_stats['FIP'] = pitching_stats['FIP'].astype(float).round(2)

    pitching_stats = pitching_stats.sort_values(by=['FIP'])
    pitching_stats = pitching_stats.reset_index(drop=True)

    for index, colname in enumerate(['Team', 'Player', 'Year', 'App', 'IP', 'W', 'L', 'SO', 'HA', 'HR A', 'WHIP', 'BAA', 'FIP']):
        cell = sheet.cell(row=len(conference_pitching_stats) + 8, column=index + 1)
        cell.value = colname
        cell.font = Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


    for index, row in pitching_stats.iterrows():
        for col in range(1, 14):
            cell = sheet.cell(row=index + len(conference_pitching_stats) + 9, column=col)
            cell.value = row[col - 1]
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            if not pitching_stats['ncaa_university_name'][index] == team_name:
                cell.fill = openpyxl.styles.PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
            elif pitching_stats['Name'][index] in players_leaving:
                cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            else:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    # Set column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 12

    returning_color_cell = sheet.cell(row=len(conference_pitching_stats) + len(pitching_stats) + 12, column=1)
    returning_color_cell.value = "Returners"
    returning_color_cell.fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    returning_color_cell.font = Font(bold=True)

    transfer_color_cell = sheet.cell(row=len(conference_pitching_stats) + len(pitching_stats) + 13, column=1)
    transfer_color_cell.value = "Transfers"
    transfer_color_cell.fill = openpyxl.styles.PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    transfer_color_cell.font = Font(bold=True)

    leaving_color_cell = sheet.cell(row=len(conference_pitching_stats) + len(pitching_stats) + 14, column=1)
    leaving_color_cell.value = "Leaving"
    leaving_color_cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    leaving_color_cell.font = Font(bold=True)

    return workbook

# Sort team names alphabetically
sorted_teams = teams['ncaa_university_name'].sort_values().tolist()

# Create a Streamlit dropdown with sorted team names
selected_team = st.selectbox("Select Team", sorted_teams)

selected_team_id = teams[teams['ncaa_university_name'] == selected_team]['ncaa_universityID'].tolist()[0]

if not pd.isnull(selected_team_id):
    query = """
        SELECT CONCAT(first_name, ' ', last_name) as Player
        FROM diamond_position_full
        WHERE cycle_id = 6 AND first_name NOT REGEXP '[0-9.]' AND ncaa_university_name = '{}'

        UNION

        SELECT CONCAT(first_name, ' ', last_name) as Player
        FROM diamond_pitching_full
        WHERE cycle_id = 6 AND first_name NOT REGEXP '[0-9.]' AND ncaa_university_name = '{}'

        ORDER BY SUBSTRING_INDEX(Player, ' ', 1);
    """.format(selected_team, selected_team)

    cursor.execute(query)

    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    team_roster = pd.DataFrame(rows, columns=columns)

    transfers_in = st.multiselect("Transfers In", all_players['Player'].tolist())
    players_leaving = st.multiselect("Players Leaving", team_roster['Player'].tolist())
    
    print(transfers_in)

    workbook = create_sheet(selected_team_id, transfers_in, players_leaving)

    with NamedTemporaryFile() as tmp:
     workbook.save(tmp.name)
     data = BytesIO(tmp.read())

    st.download_button("Download",
    data=data,
    mime='xlsx',
    file_name="{} Comp Report.xlsx".format(selected_team))

cursor.close()

connection.close()